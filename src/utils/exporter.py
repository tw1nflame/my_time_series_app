# src/utils/exporter.py
import io
import pandas as pd
from openpyxl.styles import PatternFill

def generate_excel_buffer(preds, leaderboard, static_train, ensemble_info_df):
    """
    Формирует Excel-файл в памяти с листами:
      - Predictions
      - Leaderboard с подсветкой лучшей модели
      - StaticTrainFeatures
      - WeightedEnsembleInfo
    
    Возвращает объект BytesIO с Excel-файлом.
    """
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # Лист с предсказаниями
        if preds is not None:
            preds.reset_index().to_excel(writer, sheet_name="Predictions", index=False)
        # Лидерборд с подсветкой лучшей модели
        if leaderboard is not None:
            leaderboard.to_excel(writer, sheet_name="Leaderboard", index=False)
            try:
                sheet_lb = writer.sheets["Leaderboard"]
                best_idx = leaderboard.iloc[0].name  # индекс строки лучшей модели
                fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                row_excel = best_idx + 2  # +2 из-за заголовка
                for col_idx in range(1, leaderboard.shape[1] + 1):
                    cell = sheet_lb.cell(row=row_excel, column=col_idx)
                    cell.fill = fill_green
            except Exception as e:
                print(f"Ошибка при подсветке лучшей модели в Leaderboard: {e}")
        # Лист со статическими признаками
        if static_train is not None and not static_train.empty:
            static_train.to_excel(writer, sheet_name="StaticTrainFeatures", index=False)
        # Лист с информацией об ансамбле
        if ensemble_info_df is not None and not ensemble_info_df.empty:
            ensemble_info_df.to_excel(writer, sheet_name="WeightedEnsembleInfo", index=False)
    return excel_buffer
